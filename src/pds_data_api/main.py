from fastapi import FastAPI, Request, Form, Depends, HTTPException, UploadFile, File
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse, StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session, joinedload
from typing import List
import uuid
import json
import secrets
import time
import io
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import base64
import requests
from qdrant_client import QdrantClient
from qdrant_client.http import models
import os
import uvicorn
import pathlib
from datetime import datetime
from openpyxl.styles import PatternFill

from .models import Base, Connection, Config, TableColumn, ConnectionOptions, SyncHistory
from .database import engine, get_db
from .init_db import init_db
from .pds_sync_service import PDSSyncService
from .qdrant_routes import router as qdrant_router

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,  # Changed from WARNING to DEBUG
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('pds_data_api.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Create tables if they don't exist
# Base.metadata.create_all(bind=engine)  # Commented out as tables are created in init.sql

# Initialize database with connection types
init_db()

app = FastAPI()

# Include Qdrant routes
app.include_router(qdrant_router)

# Store server start time
server_start_time = time.time()

# Get the current file's directory
current_dir = pathlib.Path(__file__).parent.resolve()

# Mount static files and templates
app.mount("/static", StaticFiles(directory=str(current_dir / "static")), name="static")
templates = Jinja2Templates(directory=str(current_dir / "templates"))

# Add json_decode filter to Jinja2 environment
def json_decode(value):
    if value:
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return None
    return None

templates.env.filters["json_decode"] = json_decode

@app.get("/")
async def root(request: Request, db: Session = Depends(get_db)):
    """Show the dashboard/home page."""
    # Get some basic stats for the dashboard
    total_connections = db.query(Connection).count()
    total_configs = db.query(Config).count()
    recent_syncs = db.query(SyncHistory).order_by(SyncHistory.created_at.desc()).limit(5).all()
    
    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "total_connections": total_connections,
            "total_configs": total_configs,
            "recent_syncs": recent_syncs
        }
    )

@app.get("/connections")
async def list_connections(request: Request, db: Session = Depends(get_db)):
    """List all connections."""
    connections = db.query(Connection).options(joinedload(Connection.connection_type)).all()
    return templates.TemplateResponse(
        "connections/list.html",
        {"request": request, "connections": connections}
    )

@app.get("/connections/new")
async def new_connection(request: Request, db: Session = Depends(get_db)):
    """Show form to create a new connection."""
    connection_types = db.query(ConnectionOptions).all()
    return templates.TemplateResponse(
        "connections/form.html",
        {"request": request, "connection": None, "connection_types": connection_types}
    )

@app.post("/connections")
async def create_connection(
    connection_name: str = Form(...),
    connection_type: str = Form(...),
    connection_config: str = Form(...),
    direction: str = Form(...),
    db: Session = Depends(get_db)
):
    """Create a new connection."""
    # Get the connection type from ConnectionOptions
    connection_type_obj = db.query(ConnectionOptions).filter(ConnectionOptions.name == connection_type).first()
    if not connection_type_obj:
        raise HTTPException(status_code=400, detail="Invalid connection type")

    connection = Connection(
        connection_name=connection_name,
        connection_type=connection_type_obj,
        connection_config=connection_config.encode(),
        direction=direction.lower() == "true"  # Convert string to boolean
    )
    db.add(connection)
    db.commit()
    return RedirectResponse(url="/connections", status_code=303)

@app.get("/connections/{connection_id}")
async def get_connection(connection_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """Get a specific connection."""
    connection = db.query(Connection).filter(Connection.id == connection_id).first()
    if not connection:
        raise HTTPException(status_code=404, detail="Connection not found")
    
    connection_types = db.query(ConnectionOptions).all()
    return templates.TemplateResponse(
        "connections/form.html",
        {"request": request, "connection": connection, "connection_types": connection_types}
    )

@app.get("/connections/{connection_id}/edit")
async def edit_connection(connection_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """Show form to edit a connection."""
    connection = db.query(Connection).filter(Connection.id == connection_id).first()
    if not connection:
        raise HTTPException(status_code=404, detail="Connection not found")
    
    connection_types = db.query(ConnectionOptions).all()
    return templates.TemplateResponse(
        "connections/form.html",
        {"request": request, "connection": connection, "connection_types": connection_types}
    )

@app.post("/connections/{connection_id}")
async def update_connection(
    connection_id: uuid.UUID,
    connection_name: str = Form(...),
    connection_type: str = Form(...),
    connection_config: str = Form(...),
    direction: str = Form(...),
    db: Session = Depends(get_db)
):
    """Update a connection."""
    connection = db.query(Connection).filter(Connection.id == connection_id).first()
    if not connection:
        raise HTTPException(status_code=404, detail="Connection not found")
    
    # Get the connection type from ConnectionOptions
    connection_type_obj = db.query(ConnectionOptions).filter(ConnectionOptions.name == connection_type).first()
    if not connection_type_obj:
        raise HTTPException(status_code=400, detail="Invalid connection type")
    
    connection.connection_name = connection_name
    connection.connection_type = connection_type_obj
    connection.connection_config = connection_config.encode()
    connection.direction = direction.lower() == "true"  # Convert string to boolean
    db.commit()
    return RedirectResponse(url="/connections", status_code=303)

@app.post("/connections/{connection_id}/delete")
async def delete_connection(connection_id: uuid.UUID, db: Session = Depends(get_db)):
    """Delete a connection."""
    connection = db.query(Connection).filter(Connection.id == connection_id).first()
    if not connection:
        raise HTTPException(status_code=404, detail="Connection not found")
    
    db.delete(connection)
    db.commit()
    return RedirectResponse(url="/connections", status_code=303)

@app.get("/pds-tables")
async def list_pds_tables(request: Request, db: Session = Depends(get_db)):
    """List all PDS tables."""
    try:
        # Query the full Config objects
        query = db.query(Config).options(
            joinedload(Config.source_connection),
            joinedload(Config.destination_connection)
        )
        print("SQL Query:", str(query.statement.compile(compile_kwargs={"literal_binds": True})))
        configs = query.all()
        print("Configs data:", [vars(config) for config in configs])
        print("First Config object attributes:", dir(configs[0]) if configs else "No configs found")
        
        # Transform configs to match template expectations
        tables = []
        for config in configs:
            tables.append({
                'id': config.id,
                'title': getattr(config, 'title', None) or config.config_name,  # Safely get title, fallback to config_name
                'table_name': config.table_name,
                'source_connection': config.source_connection,
                'destination_connection': config.destination_connection,
                'active': config.active
            })
        
        return templates.TemplateResponse(
            "pds_tables/list.html",
            {"request": request, "tables": tables}
        )
    except Exception as e:
        logger.error(f"Error in list_pds_tables: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/pds-tables/new")
async def new_pds_table(request: Request, db: Session = Depends(get_db)):
    """Show form to create a new PDS table configuration."""
    connections = db.query(Connection).all()
    return templates.TemplateResponse(
        "pds_tables/form.html",
        {"request": request, "config": None, "connections": connections}
    )

@app.post("/pds-tables")
async def create_pds_table(
    config_name: str = Form(...),
    source_connection_id: uuid.UUID = Form(...),
    destination_connection_id: uuid.UUID = Form(...),
    table_name: str = Form(...),
    title: str = Form(None),
    page_size: int = Form(1000),
    qdrant_batch_size: int = Form(100),
    db: Session = Depends(get_db)
):
    """Create a new PDS table configuration."""
    # Validate page size
    if page_size < 1 or page_size > 10000:
        raise HTTPException(status_code=400, detail="Page size must be between 1 and 10000")
    
    # Convert empty string to None for title
    title = title if title and title.strip() else None
    
    config = Config(
        config_name=config_name,
        source_connection_id=source_connection_id,
        destination_connection_id=destination_connection_id,
        table_name=table_name,
        title=title,
        page_size=page_size,
        qdrant_batch_size=qdrant_batch_size,
        active=True
    )
    db.add(config)
    db.commit()
    return RedirectResponse(url="/pds-tables", status_code=303)

@app.get("/pds-tables/{config_id}")
async def get_pds_table(config_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """Get a specific PDS table configuration."""
    config = db.query(Config).filter(Config.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    columns = db.query(TableColumn).filter(TableColumn.config_id == config_id).all()
    return templates.TemplateResponse(
        "pds_tables/columns.html",
        {"request": request, "config": config, "columns": columns}
    )

@app.get("/pds-tables/{config_id}/edit")
async def edit_pds_table(config_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """Show form to edit a PDS table configuration."""
    config = db.query(Config).filter(Config.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    connections = db.query(Connection).all()
    return templates.TemplateResponse(
        "pds_tables/form.html",
        {"request": request, "config": config, "connections": connections}
    )

@app.post("/pds-tables/{config_id}")
async def update_pds_table(
    config_id: uuid.UUID,
    config_name: str = Form(...),
    source_connection_id: uuid.UUID = Form(...),
    destination_connection_id: uuid.UUID = Form(...),
    table_name: str = Form(...),
    title: str = Form(None),
    page_size: int = Form(1000),
    qdrant_batch_size: int = Form(100),
    db: Session = Depends(get_db)
):
    """Update a PDS table configuration."""
    # Validate page size
    if page_size < 1 or page_size > 10000:
        raise HTTPException(status_code=400, detail="Page size must be between 1 and 10000")
    
    config = db.query(Config).filter(Config.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    # Convert empty string to None for title
    title = title if title and title.strip() else None
    
    config.config_name = config_name
    config.source_connection_id = source_connection_id
    config.destination_connection_id = destination_connection_id
    config.table_name = table_name
    config.title = title
    config.page_size = page_size
    config.qdrant_batch_size = qdrant_batch_size
    db.commit()
    return RedirectResponse(url="/pds-tables", status_code=303)

@app.post("/pds-tables/{config_id}/delete")
async def delete_pds_table(config_id: uuid.UUID, db: Session = Depends(get_db)):
    """Delete a PDS table configuration."""
    config = db.query(Config).filter(Config.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    db.delete(config)
    db.commit()
    return RedirectResponse(url="/pds-tables", status_code=303)

@app.get("/pds-tables/{config_id}/columns")
async def list_table_columns(config_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """List columns for a specific PDS table configuration."""
    table = db.query(Config).filter(Config.id == config_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    columns = db.query(TableColumn).filter(TableColumn.pds_table_id == config_id).all()
    return templates.TemplateResponse(
        "pds_tables/columns.html",
        {"request": request, "table": table, "columns": columns}
    )

@app.get("/pds-tables/{config_id}/columns/import-template")
async def download_import_template(config_id: uuid.UUID, db: Session = Depends(get_db)):
    """Download a template Excel file for importing columns."""
    table = db.query(Config).filter(Config.id == config_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Columns Template"
    
    # Add detailed instructions
    instructions = [
        ["Instructions for Importing Columns"],
        [""],
        ["Column Name:"],
        ["- Required field"],
        ["- Must be unique within the table"],
        ["- Should match the actual column name in your database"],
        ["- No spaces or special characters allowed"],
        [""],
        ["Data Type:"],
        ["- Required field"],
        ["- Select from the dropdown list"],
        ["- Available types: string, integer, float, boolean, date, datetime"],
        [""],
        ["Active:"],
        ["- TRUE: Column will be included in data syncs"],
        ["- FALSE: Column will be ignored during syncs"],
        [""],
        ["Primary Key:"],
        ["- TRUE: Column is part of the table's primary key"],
        ["- FALSE: Column is not part of the primary key"],
        ["- At least one column must be marked as primary key"],
        [""],
        ["Example Data:"],
        [""]
    ]
    
    # Add instructions to worksheet
    for row, instruction in enumerate(instructions, 1):
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = instruction[0]
        if row == 1:  # Title row
            cell.font = cell.font.copy(bold=True, size=14)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add headers after instructions
    header_row = len(instructions) + 2
    headers = ["Column Name", "Data Type", "Active", "Primary Key"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = cell.font.copy(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add example data
    example_data = [
        ["id", "integer", True, True],
        ["name", "string", True, False],
        ["created_at", "datetime", True, False],
        ["is_active", "boolean", True, False]
    ]
    
    for row, data in enumerate(example_data, header_row + 1):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Add data validation for Data Type column
    data_types = ["string", "integer", "float", "boolean", "date", "datetime"]
    dv = DataValidation(type="list", formula1=f'"{",".join(data_types)}"', allow_blank=True)
    dv.add(f'B{header_row + 1}:B1000')  # Apply to column B, starting after headers
    ws.add_data_validation(dv)
    
    # Add data validation for Active and Primary Key columns
    for col in ['C', 'D']:
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
        dv.add(f'{col}{header_row + 1}:{col}1000')
        ws.add_data_validation(dv)
    
    # Adjust column widths
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 20
    
    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{table.table_name}_columns_template.xlsx"'
        }
    )

@app.get("/pds-tables/{config_id}/columns/import-form")
async def import_columns_form(config_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """Show form to import columns from Excel."""
    table = db.query(Config).filter(Config.id == config_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    return templates.TemplateResponse(
        "pds_tables/import_columns.html",
        {"request": request, "table": table}
    )

@app.post("/pds-tables/{config_id}/columns/import-excel")
async def import_columns(
    config_id: uuid.UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Import columns from Excel for a specific PDS table configuration."""
    table = db.query(Config).filter(Config.id == config_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    # Read Excel file
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    # Find the header row by looking for "Column Name" in column A
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if row[0] == "Column Name":
            header_row = row_idx
            break
    
    if not header_row:
        raise HTTPException(status_code=400, detail="Invalid template format: Could not find header row")
    
    # Process rows starting after the header row
    for row in ws.iter_rows(min_row=header_row + 1):
        column_name = row[0].value
        data_type = row[1].value
        active = row[2].value
        is_primary_key = row[3].value
        
        # Skip empty rows
        if not column_name or not data_type:
            continue
            
        # Convert string 'TRUE'/'FALSE' to boolean if necessary
        active = str(active).upper() == 'TRUE' if isinstance(active, str) else bool(active)
        is_primary_key = str(is_primary_key).upper() == 'TRUE' if isinstance(is_primary_key, str) else bool(is_primary_key)
        
        column = TableColumn(
            pds_table_id=config_id,
            column_name=column_name,
            data_type=data_type,
            active=active,
            is_primary_key=is_primary_key
        )
        db.add(column)
    
    db.commit()
    return RedirectResponse(url=f"/pds-tables/{config_id}/columns", status_code=303)

@app.get("/pds-tables/{config_id}/columns/new")
async def new_table_column(config_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """Show form to create a new column for a PDS table configuration."""
    config = db.query(Config).filter(Config.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    return templates.TemplateResponse(
        "pds_tables/column_form.html",
        {"request": request, "config": config, "column": None}
    )

@app.post("/pds-tables/{config_id}/columns")
async def create_table_column(
    config_id: uuid.UUID,
    column_name: str = Form(...),
    data_type: str = Form(...),
    active: bool = Form(False),
    is_primary_key: bool = Form(False),
    db: Session = Depends(get_db)
):
    """Create a new column for a PDS table configuration."""
    column = TableColumn(
        pds_table_id=config_id,
        column_name=column_name,
        data_type=data_type,
        active=active,
        is_primary_key=is_primary_key
    )
    db.add(column)
    db.commit()
    return RedirectResponse(url=f"/pds-tables/{config_id}/columns", status_code=303)

@app.get("/pds-tables/{config_id}/columns/{column_id}/edit")
async def edit_table_column(
    config_id: uuid.UUID,
    column_id: uuid.UUID,
    request: Request,
    db: Session = Depends(get_db)
):
    """Show form to edit a column for a PDS table configuration."""
    config = db.query(Config).filter(Config.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    column = db.query(TableColumn).filter(TableColumn.id == column_id).first()
    if not column:
        raise HTTPException(status_code=404, detail="Column not found")
    
    return templates.TemplateResponse(
        "pds_tables/column_form.html",
        {"request": request, "config": config, "column": column}
    )

@app.post("/pds-tables/{config_id}/columns/{column_id}")
async def update_table_column(
    config_id: uuid.UUID,
    column_id: uuid.UUID,
    column_name: str = Form(...),
    data_type: str = Form(...),
    active: bool = Form(False),
    is_primary_key: bool = Form(False),
    db: Session = Depends(get_db)
):
    """Update a column for a PDS table configuration."""
    column = db.query(TableColumn).filter(TableColumn.id == column_id).first()
    if not column:
        raise HTTPException(status_code=404, detail="Column not found")
    
    column.column_name = column_name
    column.data_type = data_type
    column.active = active
    column.is_primary_key = is_primary_key
    db.commit()
    return RedirectResponse(url=f"/pds-tables/{config_id}/columns", status_code=303)

@app.post("/pds-tables/{config_id}/columns/{column_id}/delete")
async def delete_table_column(
    config_id: uuid.UUID,
    column_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    """Delete a column from a PDS table configuration."""
    try:
        # First verify the column exists and belongs to the correct table
        column = db.query(TableColumn).filter(
            TableColumn.id == column_id,
            TableColumn.pds_table_id == config_id
        ).first()
        
        if not column:
            raise HTTPException(status_code=404, detail="Column not found")
        
        # Delete the column
        db.delete(column)
        db.commit()
        
        # Return a redirect response
        return RedirectResponse(
            url=f"/pds-tables/{config_id}/columns",
            status_code=303
        )
    except Exception as e:
        logger.error(f"Error deleting column: {str(e)}")
        return RedirectResponse(
            url=f"/pds-tables/{config_id}/columns",
            status_code=303
        )

@app.get("/sync-history")
async def list_sync_history(request: Request, db: Session = Depends(get_db)):
    # Get all tables with their latest sync history
    tables = db.query(Config).all()
    table_history = []
    
    for table in tables:
        latest_sync = db.query(SyncHistory).filter(
            SyncHistory.pds_table_id == table.id
        ).order_by(SyncHistory.start_time.desc()).first()
        
        table_history.append({
            "table": table,
            "latest_sync": latest_sync
        })
    
    return templates.TemplateResponse(
        "sync_history/list.html",
        {
            "request": request,
            "table_history": table_history
        }
    )

@app.get("/sync-history/{table_id}")
async def get_sync_history(table_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """Get details of a specific table's sync history."""
    # Get the table details
    table = db.query(Config).filter(Config.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    
    # Get all sync history entries for this table
    sync_history = db.query(SyncHistory).filter(
        SyncHistory.pds_table_id == table_id
    ).order_by(SyncHistory.start_time.desc()).all()
    
    return templates.TemplateResponse(
        "sync_history/table.html",
        {
            "request": request,
            "table": table,
            "sync_history": sync_history
        }
    )

@app.post("/sync-history/{sync_id}/delete")
async def delete_sync_history(sync_id: uuid.UUID, db: Session = Depends(get_db)):
    """Delete a sync history entry."""
    sync_entry = db.query(SyncHistory).filter(SyncHistory.id == sync_id).first()
    if not sync_entry:
        raise HTTPException(status_code=404, detail="Sync history entry not found")
    
    db.delete(sync_entry)
    db.commit()
    return RedirectResponse(url="/sync-history", status_code=303)

@app.get("/pds-tables/{config_id}/payload")
async def view_payload(config_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """View the payload for a specific PDS table configuration."""
    table = db.query(Config).filter(Config.id == config_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    # Get the source connection details
    source_connection = table.source_connection
    if not source_connection:
        raise HTTPException(status_code=400, detail="Source connection not configured")
    
    # Get the columns for this table
    columns = db.query(TableColumn).filter(TableColumn.pds_table_id == config_id).all()
    
    # Build the PDS API request payload according to Oracle's format
    pds_payload = {
        "name": f"{table.table_name} Data",
        "pageSize": str(table.page_size),
        "tables": [
            {
                "tableName": table.table_name,
                "columns": [col.column_name for col in columns if col.active]
            }
        ]
    }
    
    return templates.TemplateResponse(
        "pds_tables/payload.html",
        {
            "request": request,
            "table": table,
            "columns": columns,
            "pds_payload": json.dumps(pds_payload, indent=2)
        }
    )

@app.post("/pds-tables/{config_id}/sync")
async def sync_table(config_id: uuid.UUID, db: Session = Depends(get_db)):
    """Trigger a sync for a specific PDS table configuration."""
    table = db.query(Config).filter(Config.id == config_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    try:
        # Initialize sync service with table_id
        sync_service = PDSSyncService(db, config_id)
        
        # Run the sync
        result = sync_service.run_sync()
        
        return RedirectResponse(url=f"/pds-tables/{config_id}/columns", status_code=303)
        
    except Exception as e:
        logger.error(f"Sync failed for table {table.table_name}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/pds-tables/{table_id}/test")
async def test_pds_connection(
    request: Request,
    table_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    try:
        # Initialize sync service
        sync_service = PDSSyncService(db, table_id)
        
        # Test the connection
        response_data = sync_service.test_connection()
        
        # Return the raw response
        return templates.TemplateResponse(
            "pds_tables/test_response.html",
            {
                "request": request,
                "table_id": table_id,
                "response": json.dumps(response_data, indent=2),
                "error": None
            }
        )
    except Exception as e:
        logger.error(f"Error testing PDS connection: {str(e)}")
        return templates.TemplateResponse(
            "pds_tables/test_response.html",
            {
                "request": request,
                "table_id": table_id,
                "response": None,
                "error": str(e)
            }
        )

@app.get("/pds-tables/{table_id}/refresh-metadata")
async def refresh_pds_metadata(
    request: Request,
    table_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    try:
        # Initialize sync service
        sync_service = PDSSyncService(db, table_id)
        
        # Get the base URL and construct the metadata refresh URL
        base_url = sync_service.source_config['url'].rstrip('/')  # Remove trailing slash if any
        if base_url.endswith('/pds'):
            base_url = base_url[:-4]  # Remove /pds from the end
        refresh_url = f"{base_url}/pds/rest-service/dataservice/metadata/refresh?configCode=ds_unifier"
        
        # Get auth header
        auth_string = f"{sync_service.source_config['username']}:{sync_service.source_config['password']}"
        encoded_auth = base64.b64encode(auth_string.encode()).decode()
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Basic {encoded_auth}"
        }
        
        logger.info(f"Making metadata refresh request to: {refresh_url}")
        
        # Make the request to refresh metadata
        response = requests.post(
            refresh_url,
            headers=headers,
            timeout=30
        )
        
        if response.status_code in [200, 202]:
            # This is a success case - the request is being processed
            logger.info("Metadata refresh initiated successfully")
            
            # Start polling for status
            status_url = f"{base_url}/pds/rest-service/v1/config/status/seed/ds_unifier"
            max_attempts = 30  # Poll for up to 30 seconds
            attempt = 0
            
            while attempt < max_attempts:
                status_response = requests.get(
                    status_url,
                    headers=headers,
                    timeout=30
                )
                
                if status_response.status_code == 200:
                    status_text = status_response.text.strip()
                    if "Metadata Seeding Status : SUCCESS" in status_text:
                        return templates.TemplateResponse(
                            "pds_tables/columns.html",
                            {
                                "request": request,
                                "table": db.query(Config).filter(Config.id == table_id).first(),
                                "flash_message": "Metadata refresh completed successfully.",
                                "flash_type": "success"
                            }
                        )
                    elif "Metadata Seeding Status : PENDING" in status_text:
                        attempt += 1
                        time.sleep(1)  # Wait 1 second before next attempt
                        continue
                
                # If we get here, something went wrong
                break
            
            # If we get here, either polling timed out or failed
            return templates.TemplateResponse(
                "pds_tables/columns.html",
                {
                    "request": request,
                    "table": db.query(Config).filter(Config.id == table_id).first(),
                    "flash_message": "Metadata refresh initiated but status check failed. Please check the logs for details.",
                    "flash_type": "warning"
                }
            )
        else:
            raise ValueError(f"Metadata refresh failed with status {response.status_code}: {response.text}")
            
    except Exception as e:
        logger.error(f"Error refreshing metadata: {str(e)}")
        return templates.TemplateResponse(
            "pds_tables/columns.html",
            {
                "request": request,
                "table": db.query(Config).filter(Config.id == table_id).first(),
                "flash_message": f"Error refreshing metadata: {str(e)}",
                "flash_type": "danger"
            }
        )

@app.get("/pds-tables/{config_id}/qdrant-view")
async def view_qdrant_data(config_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """View Qdrant data for a specific PDS table configuration."""
    table = db.query(Config).filter(Config.id == config_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    return templates.TemplateResponse(
        "pds_tables/qdrant_view.html",
        {
            "request": request,
            "table": table
        }
    )

@app.get("/pds-tables/{config_id}/columns/export")
async def export_columns(config_id: uuid.UUID, db: Session = Depends(get_db)):
    """Export columns to Excel for a specific PDS table configuration."""
    table = db.query(Config).filter(Config.id == config_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Configuration not found")
    
    columns = db.query(TableColumn).filter(TableColumn.pds_table_id == config_id).all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Columns"
    
    # Add headers
    headers = ["Column Name", "Data Type", "Active", "Primary Key"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    for row, column in enumerate(columns, 2):
        ws.cell(row=row, column=1, value=column.column_name)
        ws.cell(row=row, column=2, value=column.data_type)
        ws.cell(row=row, column=3, value=column.active)
        ws.cell(row=row, column=4, value=column.is_primary_key)
    
    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{table.table_name}_columns.xlsx"'
        }
    )

@app.post("/pds-tables/{table_id}/columns/update")
async def update_columns(
    request: Request,
    table_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    """Update multiple columns for a PDS table configuration."""
    try:
        # Log the incoming request
        logger.info(f"Received request to update columns for table {table_id}")
        
        # Get the table
        table = db.query(Config).filter(Config.id == table_id).first()
        if not table:
            logger.error(f"Table {table_id} not found")
            raise HTTPException(status_code=404, detail="Table not found")
        
        # Get form data
        form_data = await request.form()
        logger.info(f"Form data received: {dict(form_data)}")
        
        # Get all columns for this table
        columns = db.query(TableColumn).filter(TableColumn.pds_table_id == table_id).all()
        logger.info(f"Found {len(columns)} columns for table {table_id}")
        
        # Update each column's active and primary key status
        for column in columns:
            active_key = f"active_{column.id}"
            primary_key_key = f"primary_key_{column.id}"
            
            # Update active status if present in form data
            if active_key in form_data:
                column.active = form_data[active_key].lower() == "on"
            
            # Update primary key status if present in form data
            if primary_key_key in form_data:
                column.is_primary_key = form_data[primary_key_key].lower() == "on"
        
        # Commit changes
        db.commit()
        logger.info("Successfully updated columns")
        
        # Redirect back to the columns page
        return RedirectResponse(url=f"/pds-tables/{table_id}/columns", status_code=303)
        
    except Exception as e:
        logger.error(f"Error updating columns: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

def run_app(host: str = "0.0.0.0", port: int = 8000, reload: bool = False):
    """Run the FastAPI application using uvicorn."""
    uvicorn.run("pds_data_api.main:app", host=host, port=port, reload=reload)

if __name__ == "__main__":
    run_app() 