from fastapi import FastAPI, Request, Form, Depends, HTTPException, UploadFile, File
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse, StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session
from typing import List
import uuid
import json
import secrets
import time
import io
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import base64
import requests
from qdrant_client import QdrantClient
from qdrant_client.http import models
import os

from models import Base, Connection, Config, TableColumn, ConnectionOptions, SyncHistory
from database import engine, get_db
from pds_sync_service import PDSSyncService
from qdrant_routes import router as qdrant_router

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Create database tables
Base.metadata.create_all(bind=engine)

app = FastAPI()

# Include Qdrant routes
app.include_router(qdrant_router)

# Store server start time
server_start_time = time.time()

# Mount static files and templates
app.mount("/static", StaticFiles(directory="src/pds_data_api/static"), name="static")
templates = Jinja2Templates(directory="src/pds_data_api/templates")

# Add json_decode filter to Jinja2 environment
def json_decode(value):
    if value:
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return None
    return None

templates.env.filters["json_decode"] = json_decode

# Add server restart check middleware
@app.middleware("http")
async def check_server_restart(request: Request, call_next):
    response = await call_next(request)
    return response

# Add flash messages support
@app.middleware("http")
async def add_flash_messages(request: Request, call_next):
    response = await call_next(request)
    return response

# Routes for Connections
@app.get("/connections")
async def list_connections(request: Request, db: Session = Depends(get_db)):
    connections = db.query(Connection).all()
    return templates.TemplateResponse(
        "connections/list.html",
        {"request": request, "connections": connections}
    )

@app.get("/connections/create")
async def create_connection_form(request: Request, db: Session = Depends(get_db)):
    # Get connection types
    connection_types = db.query(ConnectionOptions).all()
    return templates.TemplateResponse(
        "connections/form.html",
        {"request": request, "connection": None, "connection_types": connection_types}
    )

@app.post("/connections/create")
async def create_connection(
    request: Request,
    connection_name: str = Form(...),
    connection_type: str = Form(...),
    connection_config: str = Form(...),
    direction: bool = Form(...),
    db: Session = Depends(get_db)
):
    try:
        logger.info(f"Attempting to create connection: {connection_name} of type {connection_type}")
        
        # Validate JSON config
        try:
            config_json = json.loads(connection_config)
            logger.info("JSON configuration parsed successfully")
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON configuration: {str(e)}")
            return templates.TemplateResponse(
                "connections/form.html",
                {
                    "request": request,
                    "connection": None,
                    "error": f"Invalid JSON configuration: {str(e)}",
                    "form_data": {
                        "connection_name": connection_name,
                        "connection_type": connection_type,
                        "connection_config": connection_config,
                        "direction": direction
                    }
                },
                status_code=400
            )
        
        # Validate required fields based on connection type
        if connection_type.lower() == 'pds':
            required_fields = ['url', 'username', 'password']
            missing_fields = [field for field in required_fields if field not in config_json]
            if missing_fields:
                error_msg = f"Missing required fields for PDS connection: {', '.join(missing_fields)}"
                logger.error(error_msg)
                raise ValueError(error_msg)
        elif connection_type.lower() == 'qdrant':
            required_fields = ['host', 'port']
            missing_fields = [field for field in required_fields if field not in config_json]
            if missing_fields:
                error_msg = f"Missing required fields for Qdrant connection: {', '.join(missing_fields)}"
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        # Test the connection before saving
        if connection_type.lower() == 'postgresql':
            try:
                import psycopg2
                logger.info("Attempting PostgreSQL connection...")
                # First try to connect to the database
                try:
                    conn = psycopg2.connect(
                        host=config_json.get('host'),
                        port=config_json.get('port'),
                        database=config_json.get('database'),
                        user=config_json.get('username'),
                        password=config_json.get('password')
                    )
                    conn.close()
                    logger.info("Successfully connected to existing database")
                except psycopg2.OperationalError as e:
                    if "database" in str(e) and "does not exist" in str(e):
                        logger.info("Database doesn't exist, attempting to create it...")
                        # If database doesn't exist, try to create it
                        try:
                            # Connect to postgres database to create the new database
                            conn = psycopg2.connect(
                                host=config_json.get('host'),
                                port=config_json.get('port'),
                                database='postgres',
                                user=config_json.get('username'),
                                password=config_json.get('password')
                            )
                            conn.autocommit = True
                            cursor = conn.cursor()
                            cursor.execute(f"CREATE DATABASE {config_json.get('database')}")
                            cursor.close()
                            conn.close()
                            logger.info("Successfully created new database")
                        except Exception as create_error:
                            error_msg = f"Failed to create database: {str(create_error)}"
                            logger.error(error_msg)
                            raise ValueError(error_msg)
                    else:
                        error_msg = f"Failed to connect to PostgreSQL: {str(e)}"
                        logger.error(error_msg)
                        raise ValueError(error_msg)
            except Exception as e:
                error_msg = f"Failed to connect to PostgreSQL: {str(e)}"
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        # Get or create connection type
        conn_type = db.query(ConnectionOptions).filter(ConnectionOptions.name == connection_type).first()
        if not conn_type:
            conn_type = ConnectionOptions(name=connection_type)
            db.add(conn_type)
            db.flush()
            logger.info(f"Created new connection type: {connection_type}")
        
        connection = Connection(
            connection_name=connection_name,
            connection_description=f"{connection_type} connection",
            connection_config=connection_config.encode(),  # Store as bytes
            connection_type_id=conn_type.id,
            direction=direction
        )
        db.add(connection)
        db.commit()
        logger.info(f"Successfully created connection: {connection_name}")
        
        return RedirectResponse(url="/connections", status_code=303)
    except ValueError as e:
        logger.error(f"Validation error: {str(e)}")
        return templates.TemplateResponse(
            "connections/form.html",
            {
                "request": request,
                "connection": None,
                "error": str(e),
                "form_data": {
                    "connection_name": connection_name,
                    "connection_type": connection_type,
                    "connection_config": connection_config,
                    "direction": direction
                }
            },
            status_code=400
        )
    except Exception as e:
        logger.error(f"Unexpected error creating connection: {str(e)}", exc_info=True)
        return templates.TemplateResponse(
            "connections/form.html",
            {
                "request": request,
                "connection": None,
                "error": f"An unexpected error occurred: {str(e)}",
                "form_data": {
                    "connection_name": connection_name,
                    "connection_type": connection_type,
                    "connection_config": connection_config,
                    "direction": direction
                }
            },
            status_code=500
        )

@app.get("/connections/{connection_id}/edit")
async def edit_connection_form(
    request: Request,
    connection_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    connection = db.query(Connection).filter(Connection.id == connection_id).first()
    if not connection:
        raise HTTPException(status_code=404, detail="Connection not found")
    
    # Get connection types
    connection_types = db.query(ConnectionOptions).all()
    
    return templates.TemplateResponse(
        "connections/form.html",
        {
            "request": request,
            "connection": connection,
            "connection_types": connection_types
        }
    )

@app.post("/connections/{connection_id}/edit")
async def update_connection(
    request: Request,
    connection_id: uuid.UUID,
    connection_name: str = Form(...),
    connection_type: str = Form(...),
    connection_config: str = Form(...),
    direction: bool = Form(...),
    db: Session = Depends(get_db)
):
    try:
        logger.info(f"Attempting to update connection {connection_id}")
        
        # Get connection first
        connection = db.query(Connection).filter(Connection.id == connection_id).first()
        if not connection:
            raise HTTPException(status_code=404, detail="Connection not found")
        
        # Validate JSON config
        try:
            config_json = json.loads(connection_config)
            logger.info("JSON configuration parsed successfully")
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON configuration: {str(e)}")
            return templates.TemplateResponse(
                "connections/form.html",
                {
                    "request": request,
                    "connection": connection,
                    "error": f"Invalid JSON configuration: {str(e)}",
                    "form_data": {
                        "connection_name": connection_name,
                        "connection_type": connection_type,
                        "connection_config": connection_config,
                        "direction": direction
                    }
                },
                status_code=400
            )
        
        # Validate required fields based on connection type
        if connection_type.lower() == 'pds':
            required_fields = ['url', 'username', 'password']
            missing_fields = [field for field in required_fields if field not in config_json]
            if missing_fields:
                error_msg = f"Missing required fields for PDS connection: {', '.join(missing_fields)}"
                logger.error(error_msg)
                return templates.TemplateResponse(
                    "connections/form.html",
                    {
                        "request": request,
                        "connection": connection,
                        "error": error_msg,
                        "form_data": {
                            "connection_name": connection_name,
                            "connection_type": connection_type,
                            "connection_config": connection_config,
                            "direction": direction
                        }
                    },
                    status_code=400
                )
        
        # Get or create connection type
        conn_type = db.query(ConnectionOptions).filter(ConnectionOptions.name == connection_type).first()
        if not conn_type:
            conn_type = ConnectionOptions(name=connection_type)
            db.add(conn_type)
            db.flush()
            logger.info(f"Created new connection type: {connection_type}")
        
        # Update connection
        connection.connection_name = connection_name
        connection.connection_description = f"{connection_type} connection"
        connection.connection_config = connection_config.encode()
        connection.connection_type_id = conn_type.id
        connection.direction = direction
        
        db.commit()
        logger.info(f"Successfully updated connection: {connection_name}")
        
        return RedirectResponse(url="/connections", status_code=303)
        
    except Exception as e:
        logger.error(f"Error updating connection: {str(e)}")
        return templates.TemplateResponse(
            "connections/form.html",
            {
                "request": request,
                "connection": connection if 'connection' in locals() else None,
                "error": f"An error occurred: {str(e)}",
                "form_data": {
                    "connection_name": connection_name,
                    "connection_type": connection_type,
                    "connection_config": connection_config,
                    "direction": direction
                }
            },
            status_code=500
        )

@app.post("/connections/{connection_id}/delete")
async def delete_connection(request: Request, connection_id: uuid.UUID, db: Session = Depends(get_db)):
    connection = db.query(Connection).filter(Connection.id == connection_id).first()
    if not connection:
        raise HTTPException(status_code=404, detail="Connection not found")
    
    # Check if connection is being used by any PDS tables
    source_tables = db.query(Config).filter(Config.source_connection_id == connection_id).count()
    destination_tables = db.query(Config).filter(Config.destination_connection_id == connection_id).count()
    
    if source_tables > 0 or destination_tables > 0:
        raise HTTPException(
            status_code=400,
            detail="Cannot delete connection as it is being used by one or more PDS tables"
        )
    
    db.delete(connection)
    db.commit()
    
    return RedirectResponse(url="/connections", status_code=303)

# Routes for PDS Tables
@app.get("/pds-tables")
async def list_pds_tables(request: Request, db: Session = Depends(get_db)):
    tables = db.query(Config).all()
    return templates.TemplateResponse(
        "pds_tables/list.html",
        {"request": request, "tables": tables}
    )

@app.get("/pds-tables/create")
async def create_pds_table_form(request: Request, db: Session = Depends(get_db)):
    # Get source connections (direction=True)
    source_connections = db.query(Connection).filter(
        Connection.direction == True
    ).all()
    
    # Get destination connections (direction=False)
    destination_connections = db.query(Connection).filter(
        Connection.direction == False
    ).all()
    
    return templates.TemplateResponse(
        "pds_tables/form.html",
        {
            "request": request,
            "table": None,
            "source_connections": source_connections,
            "destination_connections": destination_connections
        }
    )

@app.post("/pds-tables/create")
async def create_pds_table(
    request: Request,
    table_name: str = Form(...),
    page_size: int = Form(...),
    source_connection_id: uuid.UUID = Form(...),
    destination_connection_id: uuid.UUID = Form(...),
    title: str = Form(None),
    db: Session = Depends(get_db)
):
    try:
        # Validate page size
        if page_size < 1 or page_size > 10000:
            raise ValueError("Page size must be between 1 and 10000")
        
        table = Config(
            table_name=table_name,
            config_name=table_name,  # Set config_name to be the same as table_name
            title=title,
            source_connection_id=source_connection_id,
            destination_connection_id=destination_connection_id,
            active=True,
            page_size=page_size
        )
        db.add(table)
        db.commit()
        
        return RedirectResponse(url="/pds-tables", status_code=303)
    except ValueError as e:
        return templates.TemplateResponse(
            "pds_tables/form.html",
            {
                "request": request,
                "table": None,
                "error": str(e),
                "source_connections": db.query(Connection).filter(Connection.direction == True).all(),
                "destination_connections": db.query(Connection).filter(Connection.direction == False).all()
            },
            status_code=400
        )
    except Exception as e:
        logger.error(f"Error creating PDS table: {str(e)}")
        return templates.TemplateResponse(
            "pds_tables/form.html",
            {
                "request": request,
                "table": None,
                "error": f"An error occurred: {str(e)}",
                "source_connections": db.query(Connection).filter(Connection.direction == True).all(),
                "destination_connections": db.query(Connection).filter(Connection.direction == False).all()
            },
            status_code=500
        )

@app.get("/pds-tables/{table_id}/edit")
async def edit_pds_table_form(request: Request, table_id: uuid.UUID, db: Session = Depends(get_db)):
    table = db.query(Config).filter(Config.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="PDS Table not found")
    
    # Get source connections (direction=True)
    source_connections = db.query(Connection).filter(
        Connection.direction == True
    ).all()
    
    # Get destination connections (direction=False)
    destination_connections = db.query(Connection).filter(
        Connection.direction == False
    ).all()
    
    return templates.TemplateResponse(
        "pds_tables/form.html",
        {
            "request": request,
            "table": table,
            "source_connections": source_connections,
            "destination_connections": destination_connections
        }
    )

@app.post("/pds-tables/{table_id}/edit")
async def update_pds_table(
    request: Request,
    table_id: uuid.UUID,
    table_name: str = Form(...),
    page_size: int = Form(...),
    source_connection_id: uuid.UUID = Form(...),
    destination_connection_id: uuid.UUID = Form(...),
    title: str = Form(None),
    db: Session = Depends(get_db)
):
    try:
        # Validate page size
        if page_size < 1 or page_size > 10000:
            raise ValueError("Page size must be between 1 and 10000")
        
        table = db.query(Config).filter(Config.id == table_id).first()
        if not table:
            raise HTTPException(status_code=404, detail="PDS Table not found")
        
        table.table_name = table_name
        table.source_connection_id = source_connection_id
        table.destination_connection_id = destination_connection_id
        table.page_size = page_size
        table.title = title
        
        db.commit()
        
        return RedirectResponse(url="/pds-tables", status_code=303)
    except ValueError as e:
        return templates.TemplateResponse(
            "pds_tables/form.html",
            {
                "request": request,
                "table": table,
                "error": str(e),
                "source_connections": db.query(Connection).filter(Connection.direction == True).all(),
                "destination_connections": db.query(Connection).filter(Connection.direction == False).all()
            },
            status_code=400
        )
    except Exception as e:
        logger.error(f"Error updating PDS table: {str(e)}")
        return templates.TemplateResponse(
            "pds_tables/form.html",
            {
                "request": request,
                "table": table,
                "error": f"An error occurred: {str(e)}",
                "source_connections": db.query(Connection).filter(Connection.direction == True).all(),
                "destination_connections": db.query(Connection).filter(Connection.direction == False).all()
            },
            status_code=500
        )

@app.post("/pds-tables/{table_id}/delete")
async def delete_pds_table(request: Request, table_id: uuid.UUID, db: Session = Depends(get_db)):
    table = db.query(Config).filter(Config.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="PDS Table not found")
    
    db.delete(table)
    db.commit()
    
    return RedirectResponse(url="/pds-tables", status_code=303)

# Routes for Table Columns (now under PDS Tables)
@app.get("/pds-tables/{table_id}/columns")
async def list_table_columns(request: Request, table_id: uuid.UUID, db: Session = Depends(get_db)):
    table = db.query(Config).filter(Config.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="PDS Table not found")
    return templates.TemplateResponse(
        "pds_tables/columns.html",
        {"request": request, "table": table}
    )

@app.post("/pds-tables/{table_id}/columns/update")
async def update_columns(table_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    try:
        logger.info(f"Received update request for table_id: {table_id}")
        
        # Validate UUID format
        try:
            table_id = uuid.UUID(str(table_id))
            logger.info(f"Validated table_id UUID: {table_id}")
        except ValueError as e:
            logger.error(f"Invalid UUID format: {str(e)}")
            raise HTTPException(status_code=400, detail="Invalid table ID format")
            
        form_data = await request.form()
        logger.info(f"Received form data: {dict(form_data)}")
        
        pds_table = db.query(Config).filter(Config.id == table_id).first()
        if not pds_table:
            logger.error(f"Table not found with ID: {table_id}")
            raise HTTPException(status_code=404, detail="Table not found")
        
        logger.info(f"Found table: {pds_table.table_name} with {len(pds_table.columns)} columns")
        
        # Update all columns
        for column in pds_table.columns:
            active_key = f"active_{column.id}"
            primary_key = f"primary_key_{column.id}"
            name_key = f"name_{column.id}"
            type_key = f"type_{column.id}"
            
            logger.info(f"Processing column {column.column_name} (ID: {column.id})")
            logger.info(f"Form keys for column: active={active_key}, primary={primary_key}, name={name_key}, type={type_key}")
            
            # Update column properties if they exist in the form data
            if active_key in form_data:
                old_active = column.active
                column.active = form_data[active_key] == "on"
                logger.info(f"Updated active status: {old_active} -> {column.active}")
            
            if primary_key in form_data:
                old_primary = column.is_primary_key
                column.is_primary_key = form_data[primary_key] == "on"
                logger.info(f"Updated primary key status: {old_primary} -> {column.is_primary_key}")
            
            if name_key in form_data:
                old_name = column.column_name
                column.column_name = form_data[name_key]
                logger.info(f"Updated column name: {old_name} -> {column.column_name}")
            
            if type_key in form_data:
                old_type = column.data_type
                column.data_type = form_data[type_key]
                logger.info(f"Updated data type: {old_type} -> {column.data_type}")
        
        try:
            db.commit()
            logger.info("Successfully committed all column updates")
        except Exception as commit_error:
            logger.error(f"Error committing changes: {str(commit_error)}")
            db.rollback()
            raise
        
        return RedirectResponse(url=f"/pds-tables/{table_id}/columns", status_code=303)
    except Exception as e:
        logger.error(f"Error updating columns: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/pds-tables/{table_id}/columns/create")
async def create_table_column_form(request: Request, table_id: uuid.UUID, db: Session = Depends(get_db)):
    table = db.query(Config).filter(Config.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="PDS Table not found")
    return templates.TemplateResponse(
        "pds_tables/column_form.html",
        {"request": request, "table": table, "column": None}
    )

@app.post("/pds-tables/{table_id}/columns/create")
async def create_table_column(
    request: Request,
    table_id: uuid.UUID,
    column_name: str = Form(...),
    data_type: str = Form(...),
    active: bool = Form(False),
    is_primary_key: bool = Form(False),
    db: Session = Depends(get_db)
):
    try:
        table = db.query(Config).filter(Config.id == table_id).first()
        if not table:
            raise HTTPException(status_code=404, detail="PDS Table not found")
            
        column = TableColumn(
            pds_table_id=table_id,
            column_name=column_name,
            data_type=data_type,
            active=active,
            is_primary_key=is_primary_key
        )
        db.add(column)
        db.commit()
        
        return RedirectResponse(url=f"/pds-tables/{table_id}/columns", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/pds-tables/{table_id}/columns/create", status_code=303)

@app.get("/pds-tables/{table_id}/columns/{column_id}/edit")
async def edit_table_column_form(request: Request, table_id: uuid.UUID, column_id: uuid.UUID, db: Session = Depends(get_db)):
    table = db.query(Config).filter(Config.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="PDS Table not found")
    column = db.query(TableColumn).filter(TableColumn.id == column_id).first()
    if not column:
        raise HTTPException(status_code=404, detail="Column not found")
    return templates.TemplateResponse(
        "pds_tables/column_form.html",
        {"request": request, "table": table, "column": column}
    )

@app.post("/pds-tables/{table_id}/columns/{column_id}/edit")
async def update_table_column(
    request: Request,
    table_id: uuid.UUID,
    column_id: uuid.UUID,
    column_name: str = Form(...),
    data_type: str = Form(...),
    active: bool = Form(False),
    is_primary_key: bool = Form(False),
    db: Session = Depends(get_db)
):
    try:
        column = db.query(TableColumn).filter(TableColumn.id == column_id).first()
        if not column:
            raise HTTPException(status_code=404, detail="Column not found")
            
        column.column_name = column_name
        column.data_type = data_type
        column.active = active
        column.is_primary_key = is_primary_key
        
        db.commit()
        
        return RedirectResponse(url=f"/pds-tables/{table_id}/columns", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/pds-tables/{table_id}/columns/{column_id}/edit", status_code=303)

@app.post("/pds-tables/{table_id}/columns/{column_id}/delete")
async def delete_table_column(request: Request, table_id: uuid.UUID, column_id: uuid.UUID, db: Session = Depends(get_db)):
    column = db.query(TableColumn).filter(TableColumn.id == column_id).first()
    if not column:
        raise HTTPException(status_code=404, detail="Column not found")
    
    db.delete(column)
    db.commit()
    
    return RedirectResponse(url=f"/pds-tables/{table_id}/columns", status_code=303)

# Excel Export/Import Routes for Columns
@app.get("/pds-tables/{table_id}/columns/export")
async def export_columns(request: Request, table_id: uuid.UUID, db: Session = Depends(get_db)):
    try:
        # Get table and its columns
        table = db.query(Config).filter(Config.id == table_id).first()
        if not table:
            raise HTTPException(status_code=404, detail="PDS Table not found")
        
        columns = db.query(TableColumn).filter(TableColumn.pds_table_id == table_id).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Columns"
        
        # Add headers
        headers = ["Column Name", "Data Type", "Active"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add data
        for row, column in enumerate(columns, 2):
            ws.cell(row=row, column=1, value=column.column_name)
            ws.cell(row=row, column=2, value=column.data_type)
            ws.cell(row=row, column=3, value="Yes" if column.active else "No")
        
        # Add data validation for Active column only if there are columns
        if columns:
            dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            dv.add(f'C2:C{len(columns) + 1}')
            ws.add_data_validation(dv)
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        ws_instructions['A1'] = "Instructions for Importing Columns"
        ws_instructions['A2'] = "1. Column Name: Enter the name of the column"
        ws_instructions['A3'] = "2. Data Type: Enter the data type (e.g., VARCHAR, INTEGER, DATE)"
        ws_instructions['A4'] = "3. Active: Select 'Yes' or 'No' from the dropdown"
        ws_instructions['A6'] = "Notes:"
        ws_instructions['A7'] = "- The Active column has a dropdown with 'Yes' and 'No' options"
        ws_instructions['A8'] = "- Existing columns will be updated if they have the same name"
        ws_instructions['A9'] = "- New columns will be created if they don't exist"
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Use table_name for the filename since it's more descriptive
        filename = f"{table.table_name}_columns.xlsx"
        
        # Return Excel file
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        logger.error(f"Error exporting columns: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error exporting columns: {str(e)}")

@app.post("/pds-tables/{table_id}/columns/import")
async def import_columns(
    request: Request,
    table_id: uuid.UUID,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        # Get table
        table = db.query(Config).filter(Config.id == table_id).first()
        if not table:
            raise HTTPException(status_code=404, detail="PDS Table not found")
        
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Process rows (skip header)
        for row in ws.iter_rows(min_row=2):
            column_name = row[0].value
            data_type = row[1].value
            active_value = row[2].value
            
            # Convert active value to boolean
            active = False
            if active_value:
                active = str(active_value).lower() in ['yes', 'true', '1']
            
            if column_name and data_type:
                # Check if column exists
                existing_column = db.query(TableColumn).filter(
                    TableColumn.pds_table_id == table_id,
                    TableColumn.column_name == column_name
                ).first()
                
                if existing_column:
                    # Update existing column
                    existing_column.data_type = data_type
                    existing_column.active = active
                else:
                    # Create new column
                    column = TableColumn(
                        pds_table_id=table_id,
                        column_name=column_name,
                        data_type=data_type,
                        active=active
                    )
                    db.add(column)
        
        db.commit()
        return RedirectResponse(url=f"/pds-tables/{table_id}/columns", status_code=303)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/pds-tables/{table_id}/sync")
async def sync_pds_table(
    request: Request,
    table_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    try:
        # Initialize sync service
        sync_service = PDSSyncService(db, table_id)
        
        # Run the sync process
        response_data = sync_service.run_sync()
        
        # Set success message
        message = "Sync completed successfully"
        if response_data:
            message += f" - Received {len(response_data.get('items', []))} items"
        
        return RedirectResponse(
            url=f"/pds-tables/{table_id}/columns",
            status_code=303,
            headers={"X-Message": message}
        )
    except ValueError as ve:
        # Handle specific error messages (like rate limits)
        logger.error(f"Validation error during sync: {str(ve)}")
        return RedirectResponse(
            url=f"/pds-tables/{table_id}/columns",
            status_code=303,
            headers={"X-Message": f"Error: {str(ve)}"}
        )
    except Exception as e:
        logger.error(f"Error syncing PDS table: {str(e)}")
        return RedirectResponse(
            url=f"/pds-tables/{table_id}/columns",
            status_code=303,
            headers={"X-Message": f"Error during sync: {str(e)}"}
        )

@app.get("/pds-tables/{table_id}/payload")
async def view_pds_payload(request: Request, table_id: uuid.UUID, db: Session = Depends(get_db)):
    try:
        # Get table and its columns
        table = db.query(Config).filter(Config.id == table_id).first()
        if not table:
            raise HTTPException(status_code=404, detail="PDS Table not found")
        
        # Get active columns
        columns = db.query(TableColumn).filter(
            TableColumn.pds_table_id == table_id,
            TableColumn.active == True
        ).all()
        
        # Build the payload structure
        payload = {
            "name": table.table_name,
            "pageSize": str(table.page_size),  # Use the table's configured page size
            "tables": [
                {
                    "tableName": table.table_name,
                    "columns": [col.column_name for col in columns]
                }
            ]
        }
        
        return templates.TemplateResponse(
            "pds_tables/payload.html",
            {
                "request": request,
                "table": table,
                "payload": json.dumps(payload, indent=2)
            }
        )
    except Exception as e:
        logger.error(f"Error viewing PDS payload: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/pds-tables/{table_id}/test")
async def test_pds_connection(
    request: Request,
    table_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    try:
        # Initialize sync service
        sync_service = PDSSyncService(db, table_id)
        
        # Test the connection
        response_data = sync_service.test_connection()
        
        # Return the raw response
        return templates.TemplateResponse(
            "pds_tables/test_response.html",
            {
                "request": request,
                "table_id": table_id,
                "response": json.dumps(response_data, indent=2),
                "error": None
            }
        )
    except Exception as e:
        logger.error(f"Error testing PDS connection: {str(e)}")
        return templates.TemplateResponse(
            "pds_tables/test_response.html",
            {
                "request": request,
                "table_id": table_id,
                "response": None,
                "error": str(e)
            }
        )

@app.get("/pds-tables/{table_id}/refresh-metadata")
async def refresh_pds_metadata(
    request: Request,
    table_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    try:
        # Initialize sync service
        sync_service = PDSSyncService(db, table_id)
        
        # Get the base URL and construct the metadata refresh URL
        base_url = sync_service.source_config['url']
        refresh_url = f"{base_url}rest-service/dataservice/metadata/refresh?configCode=ds_unifier"
        
        # Get auth header
        auth_string = f"{sync_service.source_config['username']}:{sync_service.source_config['password']}"
        encoded_auth = base64.b64encode(auth_string.encode()).decode()
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Basic {encoded_auth}"
        }
        
        # Make the request to refresh metadata
        response = requests.post(
            refresh_url,
            headers=headers,
            timeout=30
        )
        
        if response.status_code == 202:
            # This is actually a success case - the request is being processed
            logger.info("Metadata refresh initiated successfully")
            return templates.TemplateResponse(
                "pds_tables/columns.html",
                {
                    "request": request,
                    "table": db.query(Config).filter(Config.id == table_id).first(),
                    "flash_message": "Metadata refresh initiated. Status will be updated automatically.",
                    "flash_type": "success"
                }
            )
        elif response.status_code != 200:
            raise ValueError(f"Metadata refresh failed with status {response.status_code}: {response.text}")
            
        # Return success message
        return templates.TemplateResponse(
            "pds_tables/columns.html",
            {
                "request": request,
                "table": db.query(Config).filter(Config.id == table_id).first(),
                "flash_message": "Metadata refresh initiated. Status will be updated automatically.",
                "flash_type": "success"
            }
        )
        
    except Exception as e:
        logger.error(f"Error refreshing metadata: {str(e)}")
        return templates.TemplateResponse(
            "pds_tables/columns.html",
            {
                "request": request,
                "table": db.query(Config).filter(Config.id == table_id).first(),
                "flash_message": f"Error refreshing metadata: {str(e)}",
                "flash_type": "danger"
            }
        )

@app.get("/pds-tables/{table_id}/poll-metadata-status")
async def poll_metadata_status(
    table_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    try:
        # Initialize sync service
        sync_service = PDSSyncService(db, table_id)
        
        # Get the base URL and construct the status URL
        base_url = sync_service.source_config['url']
        status_url = f"{base_url}rest-service/v1/config/status/seed/ds_unifier"
        
        # Get auth header
        auth_string = f"{sync_service.source_config['username']}:{sync_service.source_config['password']}"
        encoded_auth = base64.b64encode(auth_string.encode()).decode()
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Basic {encoded_auth}"
        }
        
        # Make the request to check status
        response = requests.get(
            status_url,
            headers=headers,
            timeout=30
        )
        
        if response.status_code != 200:
            raise ValueError(f"Status check failed with status {response.status_code}: {response.text}")
            
        # Parse the response - it's a plain text response
        response_text = response.text.strip()
        if "Metadata Seeding Status : SUCCESS" in response_text:
            return {"status": "SUCCESS"}
        elif "Metadata Seeding Status : PENDING" in response_text:
            return {"status": "PENDING"}
        else:
            return {"status": "UNKNOWN", "message": response_text}
        
    except Exception as e:
        logger.error(f"Error checking metadata status: {str(e)}")
        return {"status": "ERROR", "message": str(e)}

# Root route
@app.get("/")
async def root(request: Request, db: Session = Depends(get_db)):
    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "pds_tables_count": db.query(Config).count(),
            "connections_count": db.query(Connection).count(),
            "sync_history_count": db.query(SyncHistory).count()
        }
    )

@app.get("/sync-history")
async def list_sync_history(request: Request, db: Session = Depends(get_db)):
    # Get all tables with their latest sync history
    tables = db.query(Config).all()
    table_history = []
    
    for table in tables:
        latest_sync = db.query(SyncHistory).filter(
            SyncHistory.pds_table_id == table.id
        ).order_by(SyncHistory.start_time.desc()).first()
        
        table_history.append({
            "table": table,
            "latest_sync": latest_sync
        })
    
    return templates.TemplateResponse(
        "sync_history/list.html",
        {
            "request": request,
            "table_history": table_history
        }
    )

@app.get("/sync-history/{table_id}")
async def view_table_sync_history(
    request: Request,
    table_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    # Get table details
    table = db.query(Config).filter(Config.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    
    # Get sync history for this table
    sync_history = db.query(SyncHistory).filter(
        SyncHistory.pds_table_id == table_id
    ).order_by(SyncHistory.start_time.desc()).all()
    
    return templates.TemplateResponse(
        "sync_history.html",
        {
            "request": request,
            "table": table,
            "sync_history": sync_history
        }
    )

@app.post("/sync-history/clear")
async def clear_sync_history(db: Session = Depends(get_db)):
    try:
        db.query(SyncHistory).delete()
        db.commit()
        return {"status": "success", "message": "Sync history cleared successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.post("/pds-tables/{table_id}/copy")
async def copy_pds_table(request: Request, table_id: uuid.UUID, db: Session = Depends(get_db)):
    try:
        # Fetch the original table
        original_table = db.query(Config).filter(Config.id == table_id).first()
        if not original_table:
            raise HTTPException(status_code=404, detail="PDS Table not found")

        # Create a new table config (with a new name)
        new_table = Config(
            table_name=f"Copy of {original_table.table_name}",
            config_name=f"Copy of {original_table.config_name}",
            source_connection_id=original_table.source_connection_id,
            destination_connection_id=original_table.destination_connection_id,
            active=original_table.active,
            page_size=original_table.page_size
        )
        db.add(new_table)
        db.flush()  # Get new_table.id

        # Copy columns
        original_columns = db.query(TableColumn).filter(TableColumn.pds_table_id == table_id).all()
        for col in original_columns:
            new_col = TableColumn(
                pds_table_id=new_table.id,
                column_name=col.column_name,
                data_type=col.data_type,
                active=col.active,
                is_primary_key=col.is_primary_key
            )
            db.add(new_col)

        db.commit()
        # Redirect to tables list with a flash message
        response = RedirectResponse(url="/pds-tables", status_code=303)
        response.set_cookie("flash_message", "Table copied successfully!", max_age=3)
        response.set_cookie("flash_type", "success", max_age=3)
        return response
    except Exception as e:
        logger.error(f"Error copying table: {str(e)}")
        response = RedirectResponse(url="/pds-tables", status_code=303)
        response.set_cookie("flash_message", f"Error copying table: {str(e)}", max_age=3)
        response.set_cookie("flash_type", "danger", max_age=3)
        return response

@app.get("/pds-tables/{table_id}/qdrant-view", response_class=HTMLResponse)
async def view_qdrant_data(
    table_id: uuid.UUID,
    request: Request,
    db: Session = Depends(get_db),
    limit: int = 10,
    offset: int = 0
):
    """View Qdrant data for a table with Qdrant destination."""
    # Get table configuration
    table = db.query(Config).filter(Config.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    
    # Get destination connection
    dest_connection = db.query(Connection).filter(
        Connection.id == table.destination_connection_id
    ).first()
    
    if not dest_connection or dest_connection.connection_type.name.lower() != "qdrant":
        raise HTTPException(status_code=400, detail="Table is not configured with Qdrant destination")
    
    # Parse destination config
    dest_config = json.loads(dest_connection.connection_config.decode())
    
    # Initialize Qdrant client with only essential parameters
    host = dest_config.get('host', 'localhost')
    port = dest_config.get('port', 6333)
    api_key = dest_config.get('api_key')
    
    # Construct URL
    url = f"http://{host}:{port}"
    
    # Initialize with URL and API key only
    qdrant_client = QdrantClient(
        url=url,
        api_key=api_key
    )
    
    # Get collection info
    collection_name = table.table_name.lower()
    try:
        collection_info = qdrant_client.get_collection(collection_name)
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"Collection not found: {str(e)}")
    
    # Get points
    points = qdrant_client.scroll(
        collection_name=collection_name,
        limit=limit,
        offset=offset
    )
    
    return templates.TemplateResponse(
        "pds_tables/qdrant_view.html",
        {
            "request": request,
            "table": table,
            "collection_info": collection_info,
            "points": points[0],  # points[0] contains the actual points
            "limit": limit,
            "offset": offset
        }
    )

@app.post("/tables/{table_id}")
async def update_table(table_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    form_data = await request.form()
    
    table = db.query(Config).filter(Config.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")
    
    # Update table fields
    table.table_name = form_data.get("table_name")
    table.source_connection_id = uuid.UUID(form_data.get("source_connection_id"))
    table.destination_connection_id = uuid.UUID(form_data.get("destination_connection_id"))
    table.page_size = int(form_data.get("page_size"))
    table.active = form_data.get("active") == "on"
    
    # Update Qdrant batch size if destination is Qdrant
    if table.destination_connection and table.destination_connection.connection_type.name.lower() == "qdrant":
        table.qdrant_batch_size = int(form_data.get("qdrant_batch_size", 100))
    
    db.commit()
    
    return HTMLResponse(render_table_edit_form(table_id, db))

@app.get("/tables/{table_id}/destination_changed")
async def destination_changed(
    request: Request,
    table_id: uuid.UUID,
    destination_connection_id: uuid.UUID,
    db: Session = Depends(get_db)
):
    # Get the destination connection
    destination_connection = db.query(Connection).filter(Connection.id == destination_connection_id).first()
    
    # Get the table if it exists
    table = db.query(Config).filter(Config.id == table_id).first()
    
    # Check if the destination is Qdrant
    is_qdrant = destination_connection and destination_connection.connection_type.name.lower() == "qdrant"
    
    # Return the Qdrant batch size field if needed
    if is_qdrant:
        return templates.TemplateResponse(
            "pds_tables/qdrant_batch_size.html",
            {
                "request": request,
                "table": table,
                "qdrant_batch_size": table.qdrant_batch_size if table else 100
            }
        )
    return "" 